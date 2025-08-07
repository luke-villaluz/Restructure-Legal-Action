"""Generate Excel spreadsheet with contract analysis results."""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import Dict
from config.settings import SUMMARY_PATH
from utils.logger import logger

class ExcelGenerator:
    """Generate Excel spreadsheet with streamlined contract analysis."""
    
    def __init__(self):
        self.logger = logger
        self.output_dir = SUMMARY_PATH
    
    def create_blank_spreadsheet(self) -> str:
        """Create a blank Excel spreadsheet with streamlined headers."""
        os.makedirs(self.output_dir, exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Contract Analysis"
        
        self._setup_streamlined_headers(ws)
        self._adjust_column_widths(ws)
        
        filename = input("Enter the name of the file: ") + ".xlsx"
        print(f"File name set: {filename}")
        filepath = os.path.join(self.output_dir, filename)
        
        wb.save(filepath)
        self.logger.info(f"Excel spreadsheet created: {filepath}")
        return filepath

    def add_company_row(self, excel_filepath: str, company_data: Dict[str, str], row_number: int):
        """Add a single company row to the existing Excel spreadsheet."""
        wb = load_workbook(excel_filepath)
        ws = wb.active
        
        # Map company data to columns
        column_mappings = [
            ('A', 'company'),
            ('B', 'contract_name'),
            ('C', 'effective_date'),
            ('D', 'renewal_termination_date'),
            ('E', 'assignment_clause_reference'),
            ('F', 'notices_clause_present'),
            ('G', 'action_required'),
            ('H', 'recommended_action'),
            ('I', 'contact_listed')
        ]
        
        for col, field in column_mappings:
            value = company_data.get(field, 'Not Specified')
            ws[f'{col}{row_number}'] = value
            ws[f'{col}{row_number}'].alignment = Alignment(wrap_text=True, vertical="top")
        
        wb.save(excel_filepath)
        self.logger.info(f"Added row {row_number} for {company_data.get('company')}")

    def _setup_streamlined_headers(self, ws):
        """Set up the streamlined header row with formatting."""
        headers = [
            "Company",
            "Contract Name",
            "Effective Date",
            "Renewal/Termination Date",
            "Assignment Clause Reference",
            "Notices Clause Present?",
            "Action Required Prior to Name Change or Corporate Restructure",
            "Recommended Action",
            "Contact Listed"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths for streamlined columns."""
        column_widths = [20, 30, 15, 20, 35, 25, 40, 30, 25]
        
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width